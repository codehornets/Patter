#!/usr/bin/env python3
"""Docs feature drift checker.

Cross-references:
  1. The feature inventory (patter_sdk_features.xlsx) — source of truth for
     what the SDK publicly ships.
  2. The Mintlify documentation (docs/) — what the users can read about.
  3. The SDK public surface (libraries/python/getpatter/__init__.py, libraries/typescript/src/index.ts) —
     what is actually exported from the code today.

Exits with code 1 when drift exists (so the GitHub Action can open an issue);
writes a Markdown report to --output.

Usage:
    python scripts/check_feature_docs_drift.py \\
        --xlsx /path/to/patter_sdk_features.xlsx \\
        --docs docs/ \\
        --py-init libraries/python/getpatter/__init__.py \\
        --ts-index libraries/typescript/src/index.ts \\
        --output drift-report.md
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path


def _load_inventory(xlsx_path: Path) -> list[dict[str, str]]:
    """Read the feature inventory xlsx into a list of row dicts.

    Returns empty list with a clear warning on the first line of the report
    if the file is missing — the workflow falls back to a docs-only check
    in that case.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ModuleNotFoundError:
        print("openpyxl not installed; cannot read xlsx", file=sys.stderr)
        return []

    if not xlsx_path.exists():
        print(f"inventory file not found at {xlsx_path}", file=sys.stderr)
        return []

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        entry = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, row)}
        if entry.get("feature_name"):
            result.append(entry)
    return result


def _collect_docs_features(docs_dir: Path) -> set[str]:
    """Collect feature identifiers from docs/ by inspecting mdx filenames.

    Docs filenames are the canonical mapping key (e.g. silero_vad.mdx ->
    ``silero_vad``). We accept both snake_case and kebab-case.
    """
    ids: set[str] = set()
    for mdx in docs_dir.rglob("*.mdx"):
        stem = mdx.stem.lower().replace("-", "_")
        ids.add(stem)
    return ids


def _collect_sdk_exports(py_init: Path, ts_index: Path) -> set[str]:
    """Parse public exports from both SDKs — best-effort regex-based."""
    import re

    ids: set[str] = set()

    if py_init.exists():
        text = py_init.read_text()
        # __all__ entries
        all_match = re.search(r"__all__\s*=\s*\[([^\]]+)\]", text, re.S)
        if all_match:
            for token in re.findall(r"['\"]([^'\"]+)['\"]", all_match.group(1)):
                ids.add(token.lower())
        # `from X import Y, Z` lines
        for line in text.splitlines():
            m = re.match(r"\s*from\s+\S+\s+import\s+(.+)", line)
            if m:
                for token in re.split(r"[,\s]+", m.group(1)):
                    t = token.strip().rstrip(",").lower()
                    if t and not t.startswith("_"):
                        ids.add(t)

    if ts_index.exists():
        text = ts_index.read_text()
        # export { Foo, Bar } / export class Foo / export function bar
        for match in re.finditer(r"export\s+(?:class|function|const|interface|type)\s+(\w+)", text):
            ids.add(match.group(1).lower())
        for match in re.finditer(r"export\s*{\s*([^}]+)\s*}", text):
            for token in re.split(r"[,\s]+", match.group(1)):
                t = token.split(" as ")[0].strip().lower()
                if t:
                    ids.add(t)
    return ids


def _render_report(
    inventory_missing_docs: list[dict[str, str]],
    docs_missing_inventory: set[str],
    exports_missing_inventory: set[str],
    inventory_empty: bool,
) -> str:
    lines = ["# Docs feature drift report", ""]
    if inventory_empty:
        lines += [
            "⚠️ **Feature inventory not found** — cannot enforce xlsx ↔ docs "
            "parity. Check the `patter-assets` repo is reachable and the "
            "workflow secret `PATTER_ASSETS_TOKEN` is configured.",
            "",
        ]
    if inventory_missing_docs:
        lines += [
            "## Features in inventory but missing from docs",
            "",
            "| feature | status | sdk | ships_in_version |",
            "|---------|--------|-----|------------------|",
        ]
        for row in inventory_missing_docs:
            lines.append(
                f"| {row.get('feature_name', '?')} | "
                f"{row.get('status', '?')} | "
                f"{row.get('sdk', '?')} | "
                f"{row.get('ships_in_version', '?')} |"
            )
        lines.append("")
    if docs_missing_inventory:
        lines += [
            "## Docs pages without matching inventory row",
            "",
            *[f"- `{name}`" for name in sorted(docs_missing_inventory)],
            "",
        ]
    if exports_missing_inventory:
        lines += [
            "## Public SDK exports without matching inventory row",
            "",
            *[f"- `{name}`" for name in sorted(exports_missing_inventory)],
            "",
        ]
    lines += [
        "---",
        "*Generated by `scripts/check_feature_docs_drift.py`. "
        "Dispatch the `docs-sync` agent to resolve.*",
    ]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--docs", type=Path, required=True)
    parser.add_argument("--py-init", type=Path, required=True)
    parser.add_argument("--ts-index", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    inventory = _load_inventory(args.xlsx)
    inventory_ids = {row.get("feature_name", "").lower() for row in inventory if row.get("feature_name")}
    docs_ids = _collect_docs_features(args.docs)
    export_ids = _collect_sdk_exports(args.py_init, args.ts_index)

    # Inventory rows with docs_page empty OR feature_name not reachable by stem
    inventory_missing_docs = [
        row
        for row in inventory
        if row.get("feature_name", "").lower() not in docs_ids
        and row.get("status") not in ("deprecated", "removed")
    ]
    docs_missing_inventory = {
        name
        for name in docs_ids
        if name not in inventory_ids
        and name not in {"index", "introduction", "quickstart", "changelog"}
    }
    # Heuristic: exports that look like features (>3 chars, not already standard)
    standard = {"patter", "agent", "callcontrol", "patterror", "logger", "version"}
    exports_missing_inventory = {
        name
        for name in export_ids
        if name not in inventory_ids and len(name) > 3 and name not in standard
    }

    report = _render_report(
        inventory_missing_docs,
        docs_missing_inventory,
        exports_missing_inventory,
        inventory_empty=not inventory,
    )
    args.output.write_text(report)

    drifted = bool(inventory_missing_docs or docs_missing_inventory)
    print(report)
    return 1 if drifted else 0


if __name__ == "__main__":
    sys.exit(main())
